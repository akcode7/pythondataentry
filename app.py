from flask import Flask, render_template, request, redirect, url_for, flash
import openpyxl
from pathlib import Path
import os
from openpyxl.utils import get_column_letter, column_index_from_string

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'  # Required for flash messages

# Create data directory if it doesn't exist
data_dir = Path("data")
data_dir.mkdir(exist_ok=True)

# Default Excel file path
EXCEL_FILE = "./data/RN2937-OLD223-1969-INT-P-suraj.xlsx"

def ensure_excel_file_exists():
    if not os.path.exists(EXCEL_FILE):
        workbook = openpyxl.Workbook()
        workbook.save(EXCEL_FILE)

@app.route('/')
def index():
    ensure_excel_file_exists()
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    if request.method == 'POST':
        try:
            # Get form data
            names = request.form['names']
            starting_row = int(request.form['row'])
            column_letter = request.form['column'].upper()
            
            # Debug print raw input
            print(f"Raw input: {names}")
            
            # Split names and clean whitespace
            name_list = [name.strip() for name in names.split(',') if name.strip()]
            
            # Debug print parsed names
            print(f"Parsed names: {name_list}")
            print(f"Number of names: {len(name_list)}")
            
            # Convert column letter to number
            column = column_index_from_string(column_letter)
            
            # Open Excel file
            workbook = openpyxl.load_workbook(EXCEL_FILE)
            sheet = workbook.active
            
            # Write each name to the same column but different rows
            for idx, name in enumerate(name_list):
                current_row = starting_row + idx
                
                # Debug print each write operation
                print(f"Writing '{name}' to cell {column_letter}{current_row}")
                
                # Write to Excel
                sheet.cell(row=current_row, column=column).value = name
            
            # Save changes
            workbook.save(EXCEL_FILE)
            
            # Create success message
            end_row = starting_row + len(name_list) - 1
            message = f'Added {len(name_list)} names from {column_letter}{starting_row} to {column_letter}{end_row}'
            flash(message, 'success')
            print(message)  # Debug print success message
            
            return redirect(url_for('index'))
            
        except ValueError as e:
            error_msg = f'Error: Please enter valid data. Details: {str(e)}'
            flash(error_msg, 'error')
            return redirect(url_for('index'))
            
        except Exception as e:
            error_msg = f'Unexpected error: {str(e)}'
            flash(error_msg, 'error')
            return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)